#!/usr/bin/env python3
"""매핑 JSON → 엑셀 리포트.

LLM이 만든 structured_output을 받아 3개 시트로 편다.
  매핑테이블 / 미매핑API / 미매핑테스트
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CONF_FILL = {
    "상": PatternFill("solid", fgColor="C6EFCE"),
    "중": PatternFill("solid", fgColor="FFEB9C"),
    "하": PatternFill("solid", fgColor="FFC7CE"),
}

MAPPING_COLS = [
    ("api_id", "API ID", 18),
    ("api_method", "메서드", 10),
    ("api_path", "경로", 34),
    ("api_summary", "API 요약", 34),
    ("test_id", "테스트 ID", 16),
    ("test_title", "테스트 항목", 40),
    ("confidence", "일치도", 9),
    ("rationale", "판단 근거", 60),
    ("guideline_refs", "적용 지침", 22),
    ("spec_anchor", "사양서 위치", 20),
    ("test_anchor", "테스트 위치", 20),
]


def style_header(ws: Worksheet, cols: list[tuple[str, str, int]]) -> None:
    ws.append([label for _, label, _ in cols])
    for i, (_, _, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def flatten(data: dict) -> list[dict]:
    """API×테스트 카테시안 전개. 다대다는 여기서 행으로 풀린다."""
    rows: list[dict] = []
    for api in data.get("mappings", []):
        tests = api.get("tests") or [{}]
        for t in tests:
            refs = t.get("guideline_refs") or []
            rows.append(
                {
                    "api_id": api.get("api_id", ""),
                    "api_method": api.get("api_method", ""),
                    "api_path": api.get("api_path", ""),
                    "api_summary": api.get("api_summary", ""),
                    "spec_anchor": api.get("spec_anchor", ""),
                    "test_id": t.get("test_id", "(미대응)"),
                    "test_title": t.get("test_title", ""),
                    "test_anchor": t.get("test_anchor", ""),
                    "confidence": t.get("confidence", ""),
                    "rationale": t.get("rationale", ""),
                    "guideline_refs": ", ".join(refs),
                }
            )
    # 검토 우선순위: 하 → 중 → 상
    order = {"하": 0, "중": 1, "상": 2}
    rows.sort(key=lambda r: (order.get(r["confidence"], 3), r["api_path"]))
    return rows


def build(data: dict, out_path: Path) -> dict:
    wb = Workbook()

    ws = wb.active
    ws.title = "매핑테이블"
    style_header(ws, MAPPING_COLS)
    rows = flatten(data)
    conf_idx = [k for k, _, _ in MAPPING_COLS].index("confidence") + 1
    for r in rows:
        ws.append([r[k] for k, _, _ in MAPPING_COLS])
        cell = ws.cell(row=ws.max_row, column=conf_idx)
        cell.alignment = Alignment(horizontal="center")
        if fill := CONF_FILL.get(r["confidence"]):
            cell.fill = fill
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top", wrap_text=cell.column_letter in ("D", "F", "H")
            )
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("미매핑API")
    style_header(ws2, [("api_id", "API ID", 18), ("api_method", "메서드", 10),
                       ("api_path", "경로", 34), ("reason", "사유", 60)])
    for a in data.get("unmapped_apis", []):
        ws2.append([a.get("api_id", ""), a.get("api_method", ""),
                    a.get("api_path", ""), a.get("reason", "")])

    ws3 = wb.create_sheet("미매핑테스트")
    style_header(ws3, [("test_id", "테스트 ID", 16),
                       ("test_title", "테스트 항목", 40), ("reason", "사유", 60)])
    for t in data.get("unmapped_tests", []):
        ws3.append([t.get("test_id", ""), t.get("test_title", ""), t.get("reason", "")])

    ws4 = wb.create_sheet("요약")
    counts = {g: sum(1 for r in rows if r["confidence"] == g) for g in ("상", "중", "하")}
    ws4.append(["기기명", data.get("device", "")])
    ws4.append(["날짜", data.get("date", "")])
    ws4.append(["매핑 행 수", len(rows)])
    for g, c in counts.items():
        ws4.append([f"일치도 {g}", c])
    ws4.append(["미매핑 API", len(data.get("unmapped_apis", []))])
    ws4.append(["미매핑 테스트", len(data.get("unmapped_tests", []))])
    ws4.append(["적용 지침 요약", data.get("guideline_summary", "")])
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 70

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"rows": len(rows), **counts}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--output", required=True, type=Path)
    args = ap.parse_args()

    data = json.loads(args.input.read_text())
    if not data.get("mappings") and not data.get("unmapped_apis"):
        print("[excel] 매핑 결과가 비어 있음 — LLM 응답 확인 필요", file=sys.stderr)
        return 2
    stats = build(data, args.output)
    print(f"[excel] 저장 {args.output} · {stats}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
